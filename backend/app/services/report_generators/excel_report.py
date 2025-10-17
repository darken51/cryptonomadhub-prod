"""
Excel Report Generator

Generate multi-sheet Excel workbooks with formulas.
"""

from typing import Dict, Optional
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .base_report import BaseReportGenerator

logger = logging.getLogger(__name__)


class ExcelReportGenerator(BaseReportGenerator):
    """
    Excel Report Generator

    Generates Excel workbooks with multiple sheets:
    - Summary
    - Short-Term Transactions
    - Long-Term Transactions
    - Wash Sales
    """

    async def generate(self, db, data: Optional[Dict] = None) -> bytes:
        """Generate Excel workbook"""
        if data is None:
            data = await self.fetch_report_data(db)

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, data)
        self._create_transactions_sheet(wb, "Short-Term", data['short_term'])
        self._create_transactions_sheet(wb, "Long-Term", data['long_term'])

        if data['wash_sales']:
            self._create_wash_sales_sheet(wb, data['wash_sales'])

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        excel_content = output.getvalue()
        output.close()

        return excel_content

    def _create_summary_sheet(self, wb: Workbook, data: Dict):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary")
        summary = data['summary']

        # Header
        ws['A1'] = f"Cryptocurrency Tax Report - {self.tax_year}"
        ws['A1'].font = Font(size=16, bold=True, color="1e40af")
        ws.merge_cells('A1:B1')

        # Summary data
        rows = [
            ["", ""],
            ["Total Transactions", summary['total_disposals']],
            ["Short-Term Transactions", summary['short_term_transactions']],
            ["Long-Term Transactions", summary['long_term_transactions']],
            ["", ""],
            ["Short-Term Gains", summary['short_term_gain']],
            ["Short-Term Losses", -summary['short_term_loss']],
            ["Short-Term Net", summary['short_term_net']],
            ["", ""],
            ["Long-Term Gains", summary['long_term_gain']],
            ["Long-Term Losses", -summary['long_term_loss']],
            ["Long-Term Net", summary['long_term_net']],
            ["", ""],
            ["Net Capital Gain/Loss", summary['net_gain_loss']],
        ]

        for idx, row in enumerate(rows, start=2):
            ws[f'A{idx}'] = row[0]
            ws[f'B{idx}'] = row[1]

            if idx == len(rows) + 1:  # Last row
                ws[f'A{idx}'].font = Font(bold=True)
                ws[f'B{idx}'].font = Font(bold=True, size=14)

        # Format currency columns
        for row in range(6, 15):
            ws[f'B{row}'].number_format = '$#,##0.00'

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_transactions_sheet(self, wb: Workbook, sheet_name: str, transactions: list):
        """Create transactions sheet"""
        ws = wb.create_sheet(sheet_name)

        # Header
        headers = ["Asset", "Acquired", "Disposed", "Days", "Amount", "Cost Basis", "Proceeds", "Gain/Loss"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, tx in enumerate(transactions, start=2):
            ws.cell(row=row_idx, column=1, value=tx['token'])
            ws.cell(row=row_idx, column=2, value=self.format_date(tx['acquisition_date']))
            ws.cell(row=row_idx, column=3, value=self.format_date(tx['disposal_date']))
            ws.cell(row=row_idx, column=4, value=tx['holding_period_days'])
            ws.cell(row=row_idx, column=5, value=tx['amount'])
            ws.cell(row=row_idx, column=6, value=tx['cost_basis'])
            ws.cell(row=row_idx, column=7, value=tx['proceeds'])
            ws.cell(row=row_idx, column=8, value=tx['gain_loss'])

            # Format currency
            for col in [6, 7, 8]:
                ws.cell(row=row_idx, column=col).number_format = '$#,##0.00'

        # Auto-width columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_wash_sales_sheet(self, wb: Workbook, wash_sales: list):
        """Create wash sales sheet"""
        ws = wb.create_sheet("Wash Sales")

        # Header
        headers = ["Asset", "Sale Date", "Loss Amount", "Repurchase Date", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")

        # Data rows
        for row_idx, ws_data in enumerate(wash_sales, start=2):
            ws.cell(row=row_idx, column=1, value=ws_data['token'])
            ws.cell(row=row_idx, column=2, value=self.format_date(ws_data['sale_date']))
            ws.cell(row=row_idx, column=3, value=ws_data['loss_amount'])
            ws.cell(row=row_idx, column=4, value=self.format_date(ws_data['repurchase_date']))
            ws.cell(row=row_idx, column=5, value="Disallowed")

            ws.cell(row=row_idx, column=3).number_format = '$#,##0.00'
